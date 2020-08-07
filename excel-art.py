#!/usr/bin/env python
import string
from itertools import product

import click
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import PatternFill

MAX_IMAGE_WIDTH = 100
PIXEL_DIMENSION = 12


def resize_image(img: Image) -> Image:
    """Resize the image if it's over the max dimensions we can realistically work with."""
    if (max_dimension := max(img.size)) > MAX_IMAGE_WIDTH:
        ratio = max_dimension / MAX_IMAGE_WIDTH
        width, height = int(img.width / ratio), int(img.height / ratio)

        img = img.resize((width, height))

    return img


def index_to_col_name(index: int) -> str:
    """Return the name of a column at a given index, e.g. 1 -> A1."""
    columns = list(string.ascii_uppercase) + [
        f'{letter_1}{letter_2}'
        for letter_1, letter_2
        in product(string.ascii_uppercase, string.ascii_uppercase)
    ]

    return columns[index]


@click.command()
@click.option('--path', help='The path to the image you want to convert to Excel art.')
def main(path):
    """A simple program that transforms an image to an piece of Excel art."""

    with Image.open(path) as img:
        img = resize_image(img)
        width, height = img.size

        wb = Workbook()
        sheet = wb.active

        for row in range(height):
            for col in range(width):
                r, g, b = img.getpixel((col, row))
                hex_color = f'{r:02x}{g:02x}{b:02x}'

                cell = sheet.cell(row=row + 1, column=col + 1)
                cell.fill = PatternFill(fgColor=hex_color, fill_type="solid")

                if row == 0:
                    sheet.column_dimensions[index_to_col_name(col)].width = PIXEL_DIMENSION / 6

            sheet.row_dimensions[row].height = PIXEL_DIMENSION

        wb.save(filename='out.xlsx')


if __name__ == '__main__':
    main()
